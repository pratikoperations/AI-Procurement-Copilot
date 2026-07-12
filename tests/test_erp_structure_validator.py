from io import BytesIO

from openpyxl import Workbook

from modules.erp_schema_registry import REQUIRED_SHEETS
from modules.erp_structure_validator import (
    BLOCKED,
    PASS,
    PASS_WITH_WARNINGS,
    find_blank_headers,
    find_duplicate_headers,
    validate_workbook_structure,
)
from modules.erp_workbook_loader import load_erp_workbook


def _summary(
    *,
    missing_sheet=None,
    empty_sheet=None,
    blank_header_sheet=None,
    duplicate_header_sheet=None,
    include_unknown=False,
):
    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_name in REQUIRED_SHEETS:
        if sheet_name == missing_sheet:
            continue
        sheet = workbook.create_sheet(sheet_name)
        if sheet_name == blank_header_sheet:
            headers = ["Column_A", None, "Column_C"]
        elif sheet_name == duplicate_header_sheet:
            headers = ["Supplier_ID", " supplier_id ", "Country_Code"]
        else:
            headers = ["Column_A", "Column_B"]
        sheet.append(headers)
        if sheet_name != empty_sheet:
            sheet.append(["A", "B"])

    if include_unknown:
        extra = workbook.create_sheet("Notes")
        extra.append(["Note"])
        extra.append(["Harmless extra sheet"])

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return load_erp_workbook(buffer.getvalue(), filename="validation.xlsx")


def test_valid_workbook_passes():
    result = validate_workbook_structure(_summary())

    assert result.status == PASS
    assert result.blocking_count == 0
    assert result.warning_count == 0
    assert result.missing_sheets == ()
    assert result.unknown_sheets == ()


def test_missing_required_sheet_blocks():
    result = validate_workbook_structure(_summary(missing_sheet="Receipts"))

    assert result.status == BLOCKED
    assert result.missing_sheets == ("Receipts",)
    assert any(finding.code == "MISSING_REQUIRED_SHEET" for finding in result.findings)


def test_empty_required_sheet_blocks():
    result = validate_workbook_structure(_summary(empty_sheet="Supplier_Performance"))

    assert result.status == BLOCKED
    assert any(
        finding.code == "EMPTY_REQUIRED_SHEET"
        and finding.sheet_name == "Supplier_Performance"
        for finding in result.findings
    )


def test_blank_header_blocks_and_reports_position():
    result = validate_workbook_structure(_summary(blank_header_sheet="Material_Master"))

    assert result.status == BLOCKED
    finding = next(item for item in result.findings if item.code == "BLANK_HEADER")
    assert finding.sheet_name == "Material_Master"
    assert "2" in finding.message


def test_duplicate_header_blocks_case_insensitively():
    result = validate_workbook_structure(
        _summary(duplicate_header_sheet="Supplier_Master")
    )

    assert result.status == BLOCKED
    finding = next(item for item in result.findings if item.code == "DUPLICATE_HEADER")
    assert finding.sheet_name == "Supplier_Master"
    assert "Supplier_ID" in finding.message


def test_unknown_sheet_is_logged_without_blocking():
    result = validate_workbook_structure(_summary(include_unknown=True))

    assert result.status == PASS_WITH_WARNINGS
    assert result.unknown_sheets == ("Notes",)
    assert result.blocking_count == 0
    assert any(finding.code == "UNKNOWN_SHEET" for finding in result.findings)


def test_header_helpers_are_deterministic():
    assert find_blank_headers(["A", None, " ", "D"]) == (2, 3)
    assert find_duplicate_headers(["Supplier_ID", " supplier_id ", "Country"]) == (
        "Supplier_ID",
    )


def test_adversarial_sample_is_structurally_blocked():
    summary = load_erp_workbook(
        "data/erp_samples/erp_adversarial_procurement_workbook.xlsx"
    )
    result = validate_workbook_structure(summary)

    assert result.status == BLOCKED
    assert result.blocking_count > 0
