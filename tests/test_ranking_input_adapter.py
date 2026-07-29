from __future__ import annotations

from copy import deepcopy
from datetime import date
from hashlib import sha256
from io import BytesIO
import json
from pathlib import Path

from openpyxl import Workbook

from modules.ranking_input_models import RankingMappingConfirmation
from modules.rfq_workbook_adapter import SCHEMA_PATH, adapt_v13_workbook

FIXTURE = Path("tests/fixtures/ranking_input_contract_examples.json")


def _required(definition: str):
    schema = json.loads(Path(SCHEMA_PATH).read_text(encoding="utf-8"))
    return schema["$defs"][definition]["required"]


def _value(field: str, supplier: str):
    values = {
        "SOURCING_EVENT_ID": "EVT-001", "RFQ_NUMBER": "0010000001", "RFQ_ITEM": "00010",
        "QUOTATION_VERSION": 1, "SUPPLIER_ID": supplier, "SUPPLIER_NAME": f"Supplier {supplier[-1]}",
        "MATERIAL_DESCRIPTION": "Carton", "MATERIAL_GROUP": "PACK", "PURCHASING_ORG": "1000",
        "REQUESTED_QUANTITY": 1000, "QUOTED_QUANTITY": 1000, "QUOTATION_UOM": "EA",
        "COMPARISON_UOM": "EA", "BASE_UNIT_PRICE": 12.5, "PRICE_UNIT": 1, "CURRENCY": "INR",
        "QUOTATION_DATE": "2026-07-01", "VALIDITY_END_DATE": "2026-12-31", "QUOTATION_STATUS": "VALID",
        "SOURCE_TRANSACTION": "ME49", "SOURCE_FILE_NAME": "c2.xlsx",
        "SOURCE_EXTRACTED_AT": "2026-07-29T10:00:00", "SOURCE_ROW_ID": f"RFQ-{supplier}",
    }
    return values.get(field, "SYNTHETIC")


def _metadata():
    return {
        "UPLOAD_ID": "UP-C2", "SCHEMA_VERSION": "1.3.1", "UPLOAD_MODE": "QUICK_RFQ",
        "SOURCE_SYSTEM": "SAP", "PURCHASING_ORG": "1000", "BASE_CURRENCY": "INR",
        "EXTRACTED_AT": "2026-07-29T10:00:00", "UPLOAD_CREATED_AT": "2026-07-29T10:05:00",
        "RFQ_SOURCE_TRANSACTION": "ME49", "DATA_CLASSIFICATION": "SYNTHETIC",
        "ANONYMIZATION_STATUS": "SYNTHETIC", "SOURCE_FILE_HASH_SHA256": "a" * 64,
        "NOTES": "Generated C2 fixture",
    }


def _book(alias_otif: bool = False, *, ranking_overrides=None, extra_ranking_property=False):
    wb = Workbook(); rfq = wb.active; rfq.title = "RFQ_QUOTES"
    headers = _required("RFQQuoteRow"); rfq.append(headers)
    for supplier in ("0000100001", "0000100002"):
        rfq.append([_value(field, supplier) for field in headers])
    md = wb.create_sheet("UPLOAD_METADATA"); metadata = _metadata(); md.append(list(metadata)); md.append(list(metadata.values()))
    ranking = wb.create_sheet("SUPPLIER_RANKING_INPUTS")
    template = json.loads(FIXTURE.read_text(encoding="utf-8"))["valid_quick"]
    rows = []
    for index, supplier in enumerate(("0000100001", "0000100002"), start=1):
        row = deepcopy(template)
        row["RANKING_INPUT_RECORD_ID"] = f"RANK-{index}"
        row["SUPPLIER_ID"] = supplier
        row["SUPPLIER_NAME"] = f"Supplier {index}"
        row["VALUE_ORIGINS"] = {field: "SOURCE_MAPPED" for field, value in row.items() if field in {
            "OTIF_PERCENT", "QUALITY_PPM", "SUPPLIER_AUDIT_SCORE", "COMPLAINT_RATE_PERCENT",
            "CAPACITY_BUFFER_PERCENT", "RECYCLABILITY_PERCENT", "CERTIFICATION_SCORE", "CARBON_SCORE",
            "EPR_READINESS_SCORE", "PCR_CONTENT_PERCENT",
        } and value is not None}
        for key, value in (ranking_overrides or {}).items():
            row[key] = deepcopy(value)
        if extra_ranking_property:
            row["UNCONTROLLED_SCORE"] = 99
        rows.append(row)
    ranking_headers = list(rows[0])
    if alias_otif:
        ranking_headers[ranking_headers.index("OTIF_PERCENT")] = "OTIF %"
    ranking.append(ranking_headers)
    for row in rows:
        values = []
        for header in ranking_headers:
            canonical = "OTIF_PERCENT" if header == "OTIF %" else header
            value = row.get(canonical)
            values.append(json.dumps(value) if isinstance(value, dict) else value)
        ranking.append(values)
    stream = BytesIO(); wb.save(stream); wb.close(); return stream.getvalue()


def test_v131_generated_workbook_produces_ranking_review_evidence():
    result = adapt_v13_workbook(_book(), filename="c2.xlsx", evaluation_date=date(2026, 8, 1))
    assert result.schema_version == "1.3.1"
    assert len(result.supplier_ranking_inputs) == 2
    assert len(result.ranking_evidence_results) == 20
    assert len(result.ranking_scope_matches) == 2
    assert all(item.status == "RANKING_REVIEW_COMPLETE" for item in result.ranking_mode_eligibility)
    assert any(item.code == "CANONICAL_RANKING_INPUTS_AVAILABLE_FOR_REVIEW" for item in result.findings)
    assert not any(item.code == "RANKING_ROW_SCHEMA_INVALID" for item in result.findings)


def test_noncanonical_ranking_alias_requires_origin_bound_confirmation():
    payload = _book(alias_otif=True)
    result = adapt_v13_workbook(payload, filename="c2.xlsx", evaluation_date=date(2026, 8, 1))
    review = next(item for item in result.mapping_reviews if item.source_header == "OTIF %")
    assert review.requires_confirmation
    wrong = RankingMappingConfirmation(sha256(payload).hexdigest(), "1.3.1", "1.3.1", "SUPPLIER_RANKING_INPUTS", "OTIF %", "OTIF_PERCENT", "0_TO_100_ONLY", "DERIVED_FROM_HISTORY")
    wrong_result = adapt_v13_workbook(payload, filename="c2.xlsx", evaluation_date=date(2026, 8, 1), ranking_confirmations=(wrong,))
    assert any(item.code == "RANKING_MAPPING_CONFIRMATION_REQUIRED" for item in wrong_result.findings)
    correct = RankingMappingConfirmation(sha256(payload).hexdigest(), "1.3.1", "1.3.1", "SUPPLIER_RANKING_INPUTS", "OTIF %", "OTIF_PERCENT", "0_TO_100_ONLY", "SOURCE_MAPPED")
    confirmed = adapt_v13_workbook(payload, filename="c2.xlsx", evaluation_date=date(2026, 8, 1), ranking_confirmations=(correct,))
    assert not any(item.code == "RANKING_MAPPING_CONFIRMATION_REQUIRED" and item.field_name == "OTIF_PERCENT" for item in confirmed.findings)


def test_runtime_schema_rejects_unknown_ranking_property():
    result = adapt_v13_workbook(_book(extra_ranking_property=True), filename="schema.xlsx", evaluation_date=date(2026, 8, 1))
    assert any(item.code == "RANKING_ROW_SCHEMA_INVALID" for item in result.findings)
    assert all(not record.row_valid for record in result.supplier_ranking_inputs)


def test_adapter_detects_automatic_scale_ambiguity_without_conversion():
    result = adapt_v13_workbook(_book(ranking_overrides={"OTIF_PERCENT": 0.95}), filename="scale.xlsx", evaluation_date=date(2026, 8, 1))
    otif = [item for item in result.ranking_evidence_results if item.canonical_field == "OTIF_PERCENT"]
    assert otif and all(item.canonical_evidence_status == "AMBIGUOUS_SCALE" for item in otif)
    assert all(float(item.canonical_value) == 0.95 for item in otif)


def test_adapter_rejects_engine_default_origin_and_invalid_period_controls():
    origins = {
        "OTIF_PERCENT": "DEFAULTED_BY_ENGINE", "QUALITY_PPM": "SOURCE_MAPPED",
        "SUPPLIER_AUDIT_SCORE": "SOURCE_MAPPED", "RECYCLABILITY_PERCENT": "SOURCE_MAPPED",
        "CERTIFICATION_SCORE": "SOURCE_MAPPED",
    }
    result = adapt_v13_workbook(_book(ranking_overrides={"VALUE_ORIGINS": origins}), filename="origin.xlsx", evaluation_date=date(2026, 8, 1))
    assert any(item.code == "ENGINE_DEFAULT_ORIGIN_PROHIBITED" and item.severity == "Fatal" for item in result.findings)
    period = adapt_v13_workbook(_book(ranking_overrides={"MEASUREMENT_PERIOD_START_DATE": "2026-08-02", "PERFORMANCE_RECORD_COUNT": 0}), filename="period.xlsx", evaluation_date=date(2026, 8, 1))
    assert any(item.code == "MEASUREMENT_PERIOD_INVALID" for item in period.findings)
    assert any(item.code == "PERFORMANCE_RECORD_COUNT_INVALID" for item in period.findings)
