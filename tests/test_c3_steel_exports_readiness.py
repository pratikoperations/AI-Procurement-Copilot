from io import BytesIO
import json
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
import pytest

from modules.data_loader import get_demo_data
from modules.steel_exports import (
    STEEL_EXCEL_SHEETS,
    build_steel_excel_workbook,
    build_steel_governance_manifest,
    build_steel_json_export,
    normalize_strict_steel_json,
)
from modules.steel_scenario import STEEL_SCENARIOS, run_governed_steel_scenarios
from modules.steel_ux import normalize_steel_dependent_state


def _bundle(profile="CR_COIL_COMMERCIAL", scenario="Base Case", display="Both"):
    suppliers = get_demo_data("Raw Material Procurement", "Steel")
    summary, details = run_governed_steel_scenarios(suppliers, profile, 500000, 83, display)
    state = normalize_steel_dependent_state({
        "steel_profile": profile,
        "steel_scenario": scenario,
        "steel_sourcing_route": "Domestic",
        "steel_zinc_cost_usd_per_kg": 0.0 if profile == "CR_COIL_COMMERCIAL" else 0.08,
        "steel_paint_treatment_usd_per_kg": 0.12 if profile == "PPGI_COIL_Z120" else 0.0,
        "steel_import_duty_pct": 0.0,
        "steel_substitution_status": "Non-applicable",
        "display_currency": display,
    })
    selected = details[scenario]
    manifest = build_steel_governance_manifest(state, summary, selected, 83, display)
    return state, summary, selected, manifest


def test_excel_contains_exact_governed_sheet_set_and_reconciles_visible_state():
    state, summary, selected, manifest = _bundle()
    payload = build_steel_excel_workbook(state, summary, selected, manifest)
    workbook = load_workbook(BytesIO(payload), data_only=True)
    assert tuple(workbook.sheetnames) == STEEL_EXCEL_SHEETS
    governance = workbook["C3 Governance"]
    fields = {governance.cell(row=i, column=1).value: governance.cell(row=i, column=2).value for i in range(2, governance.max_row + 1)}
    assert fields["selected_profile"] == state["steel_profile"]
    assert fields["selected_scenario"] == state["steel_scenario"]
    assert fields["winner_state"] == selected["recommendation"]["winner_state"]
    assert workbook["Scenarios"].max_row == len(STEEL_SCENARIOS) + 1


def test_strict_json_has_top_level_governance_and_separate_currency_fields():
    state, summary, selected, manifest = _bundle("PPGI_COIL_Z120")
    raw = build_steel_json_export(manifest)
    assert b"NaN" not in raw and b"Infinity" not in raw
    payload = json.loads(raw)
    governance = payload["steel_governance"]
    assert governance["selected_profile"] == "PPGI_COIL_Z120"
    assert governance["human_approval_required"] is True
    assert governance["autonomous_award"] is False
    assert governance["engineering_approval_provided"] is False
    assert governance["live_market_data_claim"] is False
    supplier = governance["supplier_scores"][0]
    assert isinstance(supplier["normalized_usd_per_kg"], (int, float))
    assert isinstance(supplier["equivalent_inr_per_kg"], (int, float))
    assert len(governance["scenarios"]) == 7


def test_non_finite_values_normalize_to_json_null():
    assert normalize_strict_steel_json({"a": float("nan"), "b": np.inf, "c": -np.inf}) == {"a": None, "b": None, "c": None}


@pytest.mark.parametrize("profile", ["CR_COIL_COMMERCIAL", "GI_COIL_Z120", "PPGI_COIL_Z120"])
def test_all_profiles_generate_reconcilable_exports(profile):
    state, summary, selected, manifest = _bundle(profile)
    governance = manifest["steel_governance"]
    assert governance["selected_profile"] == profile
    assert governance["winner_state"] == selected["recommendation"]["winner_state"]
    assert governance["optimized_unallocated_volume_kg"] == pytest.approx(selected["optimized_allocation"].attrs["unallocated_volume_kg"])
    assert len(summary) == 7


def test_changed_winner_no_winner_and_partial_allocation_are_exportable():
    state, summary, selected, manifest = _bundle(scenario="Mill Allocation and Capacity Stress")
    assert manifest["steel_governance"]["winner_state"] == selected["recommendation"]["winner_state"]
    suppliers = get_demo_data("Raw Material Procurement", "Steel")
    suppliers["Supplier Capacity"] = 1.0
    summary2, details2 = run_governed_steel_scenarios(suppliers, "CR_COIL_COMMERCIAL", 500000, 83, "Both")
    state2 = dict(state)
    state2["steel_scenario"] = "Base Case"
    manifest2 = build_steel_governance_manifest(state2, summary2, details2["Base Case"], 83, "Both")
    assert manifest2["steel_governance"]["winner"] is None
    assert manifest2["steel_governance"]["optimized_unallocated_volume_kg"] == pytest.approx(500000)


def test_steel_route_preserves_dedicated_exports_and_continues_to_common_sections():
    steel_source = Path("modules/steel_ux.py").read_text(encoding="utf-8")
    dashboard_source = Path("modules/dashboard.py").read_text(encoding="utf-8")
    assert "build_steel_excel_workbook" in steel_source
    assert "build_steel_json_export" in steel_source
    assert "st.stop()" not in steel_source
    assert "render_steel_governed_dashboard" in dashboard_source
    assert "Download Steel Decision Audit JSON" in steel_source


def test_historical_c1_c2_and_generic_export_contract_remains_present():
    exports_source = Path("modules/exports.py").read_text(encoding="utf-8")
    assert "build_c2_export_manifest" in exports_source
    assert "build_excel_workbook" in exports_source
    assert "build_decision_package_json" in exports_source
