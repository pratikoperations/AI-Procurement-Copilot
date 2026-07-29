from hashlib import sha256
from io import BytesIO
import json
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook

from modules import rfq_integration_controller as controller
from modules.rfq_review_state import ReviewState
from modules.rfq_workbook_adapter import SCHEMA_PATH


def test_route_flag_is_disabled_by_default():
    enabled, warning = controller.governed_route_enabled({})
    assert not enabled and warning is None


def test_malformed_route_flag_fails_closed():
    enabled, warning = controller.governed_route_enabled({controller.ROUTE_FLAG: "maybe"})
    assert not enabled and controller.ROUTE_FLAG in warning


def test_no_file_returns_no_dataframe():
    result = controller.run_governed_review(None, env={controller.ROUTE_FLAG: "true"}, session_state={})
    assert result.review_state is ReviewState.NO_FILE
    assert result.dataframe is None and not result.analysis_handoff_allowed


def test_upload_hash_reset_preserves_same_file_and_clears_changed_file():
    state = {
        "governed_v13_confirmed_mappings": (("RFQ_QUOTES", "RFQ-Number", "RFQ_NUMBER"),),
        "governed_v13_selected_event": "EVT-001",
    }
    assert controller.reset_session_for_upload(state, b"one")
    state["governed_v13_selected_event"] = "EVT-001"
    assert not controller.reset_session_for_upload(state, b"one")
    assert state["governed_v13_selected_event"] == "EVT-001"
    assert controller.reset_session_for_upload(state, b"two")
    assert "governed_v13_selected_event" not in state
    assert state[controller.SESSION_UPLOAD_HASH_KEY] == sha256(b"two").hexdigest()


def test_adapter_fatal_cannot_fall_through(monkeypatch):
    fake = SimpleNamespace(findings=(SimpleNamespace(severity="Fatal", code="X"),), mapping_reviews=(), available_sourcing_event_ids=())
    monkeypatch.setattr(controller, "adapt_v13_workbook", lambda *args, **kwargs: fake)
    result = controller.run_governed_review(b"workbook", filename="x.xlsx", env={controller.ROUTE_FLAG: "true"}, session_state={})
    assert result.review_state is ReviewState.ADAPTER_FATAL and result.dataframe is None


def test_insufficient_evidence_never_reaches_compatibility(monkeypatch):
    adapter = SimpleNamespace(findings=(), mapping_reviews=(), available_sourcing_event_ids=("E1",), selected_sourcing_event_id="E1", mode="QUICK_RFQ")
    orchestration = SimpleNamespace(eligibility_status="INSUFFICIENT_EVIDENCE", conditional_findings=())
    monkeypatch.setattr(controller, "adapt_v13_workbook", lambda *args, **kwargs: adapter)
    monkeypatch.setattr(controller, "orchestrate_adapter_result", lambda *args, **kwargs: orchestration)
    called = {"compatibility": False}
    monkeypatch.setattr(controller, "assess_legacy_compatibility", lambda *args, **kwargs: called.update(compatibility=True))
    state = {controller.SESSION_UPLOAD_HASH_KEY: controller.upload_sha256(b"workbook")}
    result = controller.run_governed_review(b"workbook", filename="x.xlsx", selected_sourcing_event_id="E1", env={controller.ROUTE_FLAG: "true"}, session_state=state)
    assert result.review_state is ReviewState.INSUFFICIENT_EVIDENCE
    assert result.dataframe is None and not called["compatibility"]


def _schema():
    return json.loads(Path(SCHEMA_PATH).read_text(encoding="utf-8"))


def _required(name):
    return list(_schema()["$defs"][name]["required"])


def _value(field, supplier="0000100001", event="EVT-001", currency="INR"):
    values = {
        "SOURCING_EVENT_ID": event, "RFQ_NUMBER": "0010000001", "RFQ_ITEM": "00010",
        "QUOTATION_VERSION": 1, "SUPPLIER_ID": supplier, "SUPPLIER_NAME": f"Supplier {supplier[-1]}",
        "MATERIAL_ID": "MAT-1", "MATERIAL_DESCRIPTION": "Synthetic carton", "MATERIAL_GROUP": "PACK",
        "PURCHASING_ORG": "1000", "REQUESTED_QUANTITY": 1000, "QUOTED_QUANTITY": 1000,
        "QUOTATION_UOM": "EA", "COMPARISON_UOM": "EA", "BASE_UNIT_PRICE": 830 if currency == "INR" else 10,
        "PRICE_UNIT": 1, "CURRENCY": currency, "EXCHANGE_RATE": 83 if currency == "INR" else None,
        "EXCHANGE_RATE_DATE": "2026-07-29", "QUOTATION_DATE": "2026-07-01", "VALIDITY_END_DATE": "2026-12-31",
        "QUOTATION_STATUS": "VALID", "SOURCE_TRANSACTION": "ME49", "SOURCE_FILE_NAME": "fixture.xlsx",
        "SOURCE_EXTRACTED_AT": "2026-07-29T10:00:00", "SOURCE_ROW_ID": f"RFQ-{event}-{supplier}",
        "FULL_QUANTITY_AVAILABLE": True, "PAYMENT_TERMS_CODE": "NET30", "INCOTERMS_CODE": "DDP",
        "LEAD_TIME_DAYS": 10, "TECHNICALLY_APPROVED": True, "RISK_SCORE": 80, "ESG_SCORE": 75,
        "PO_NUMBER": "0045000001", "PO_ITEM": "00010", "PO_DATE": "2026-06-01", "ORDER_QUANTITY": 1000,
        "ORDER_UOM": "EA", "NET_PRICE": 10, "NET_ORDER_VALUE": 10000, "PO_STATUS": "COMPLETE",
        "DELETION_FLAG": False,
    }
    return values.get(field, "SYNTHETIC")


def _history_row(headers, *, material_id, po_number, po_date, source_row_id):
    overrides = {
        "MATERIAL_ID": material_id,
        "PO_NUMBER": po_number,
        "PO_DATE": po_date,
        "SOURCE_ROW_ID": source_row_id,
        "SOURCE_EXTRACTED_AT": "2026-07-15T10:00:00",
        "CURRENCY": "USD",
        "NET_PRICE": 10,
        "NET_ORDER_VALUE": 10000,
    }
    return [overrides.get(field, _value(field, currency="USD")) for field in headers]


def _fixture_workbook():
    wb = Workbook()
    rfq = wb.active
    rfq.title = "RFQ_QUOTES"
    required = _required("RFQQuoteRow")
    optional = ["FULL_QUANTITY_AVAILABLE", "PAYMENT_TERMS_CODE", "INCOTERMS_CODE", "LEAD_TIME_DAYS", "TECHNICALLY_APPROVED", "RISK_SCORE", "ESG_SCORE", "EXCHANGE_RATE", "EXCHANGE_RATE_DATE"]
    headers = ["RFQ-Number" if field == "RFQ_NUMBER" else field for field in required] + [field for field in optional if field not in required]
    rfq.append(headers)
    for event, supplier, currency in (("EVT-001", "0000100001", "INR"), ("EVT-001", "0000100002", "USD"), ("EVT-002", "0000100003", "INR"), ("EVT-002", "0000100004", "USD")):
        rfq.append([_value("RFQ_NUMBER" if header == "RFQ-Number" else header, supplier, event, currency) for header in headers])

    po = wb.create_sheet("PO_HISTORY")
    po_headers = _required("POHistoryRow")
    po.append(po_headers)
    po.append(_history_row(po_headers, material_id="MAT-1", po_number="0045000001", po_date="2026-06-01", source_row_id="PO-CURRENT-MAT-1"))
    po.append(_history_row(po_headers, material_id="MAT-X", po_number="0045000002", po_date="2025-12-01", source_row_id="PO-OUTSIDE-WINDOW"))

    md = wb.create_sheet("UPLOAD_METADATA")
    md_headers = ["UPLOAD_ID", "SCHEMA_VERSION", "UPLOAD_MODE", "SOURCE_SYSTEM", "PURCHASING_ORG", "BASE_CURRENCY", "EXTRACTED_AT", "UPLOAD_CREATED_AT", "RFQ_SOURCE_TRANSACTION", "DATA_CLASSIFICATION", "ANONYMIZATION_STATUS", "SOURCE_FILE_HASH_SHA256", "HISTORY_START_DATE", "HISTORY_END_DATE", "HISTORY_SOURCE_TRANSACTION", "NOTES"]
    md.append(md_headers)
    md.append(["UP-001", "1.3.0", "FULL_SOURCING_REVIEW", "SAP", "1000", "USD", "2026-07-29T10:00:00", "2026-07-29T10:05:00", "ME49", "SYNTHETIC", "SYNTHETIC", "a" * 64, "2026-01-01", "2026-07-29", "ME2N", "Integration fixture"])
    stream = BytesIO()
    wb.save(stream)
    wb.close()
    return stream.getvalue()


def test_full_workbook_flow_is_review_only_and_preserves_mixed_currency():
    payload = _fixture_workbook()
    env = {controller.ROUTE_FLAG: "true"}
    session = {}
    first = controller.run_governed_review(payload, filename="fixture.xlsx", env=env, session_state=session)
    assert first.review_state is ReviewState.MAPPING_CONFIRMATION_REQUIRED
    confirmed = (("RFQ_QUOTES", "RFQ-Number", "RFQ_NUMBER"),)
    second = controller.run_governed_review(payload, filename="fixture.xlsx", confirmed_mappings=confirmed, env=env, session_state=session)
    assert second.review_state is ReviewState.EVENT_SELECTION_REQUIRED
    third = controller.run_governed_review(payload, filename="fixture.xlsx", confirmed_mappings=confirmed, selected_sourcing_event_id="EVT-001", env=env, session_state=session)
    assert third.review_state is ReviewState.ITEM_SELECTION_REQUIRED
    fourth = controller.run_governed_review(payload, filename="fixture.xlsx", confirmed_mappings=confirmed, selected_sourcing_event_id="EVT-001", selected_rfq_number="0010000001", selected_rfq_item="00010", env=env, session_state=session)
    assert fourth.review_state is ReviewState.CONDITIONAL_REVIEW_REQUIRED
    assert "HISTORY_ROW_OUT_OF_WINDOW" in fourth.orchestration_result.warnings
    fourth = controller.run_governed_review(payload, filename="fixture.xlsx", confirmed_mappings=confirmed, selected_sourcing_event_id="EVT-001", selected_rfq_number="0010000001", selected_rfq_item="00010", acknowledged_warning_codes=("HISTORY_ROW_OUT_OF_WINDOW",), env=env, session_state=session)
    assert fourth.review_state is ReviewState.REVIEW_ONLY_COMPLETE
    assert fourth.dataframe is None and not fourth.analysis_handoff_allowed
    currencies = {item.normalization.normalized_values["SOURCE_CURRENCY"] for item in fourth.orchestration_result.enriched_quotes}
    assert currencies == {"INR", "USD"}
    assert fourth.orchestration_result.comparison_currency == "USD"
