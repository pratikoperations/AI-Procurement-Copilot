from __future__ import annotations

from hashlib import sha256
from io import BytesIO
import json
from pathlib import Path

from openpyxl import Workbook
import pytest

from modules.rfq_workbook_adapter import (
    ALIAS_PATH,
    SCHEMA_PATH,
    WorkbookAdapterError,
    adapt_v13_workbook,
)


def _schema() -> dict:
    return json.loads(Path(SCHEMA_PATH).read_text(encoding="utf-8"))


def _required_headers(definition: str) -> list[str]:
    return list(_schema()["$defs"][definition]["required"])


def _value(
    field: str,
    supplier: str = "0000100001",
    event: str = "EVT-001",
    version: int = 1,
):
    values = {
        "SOURCING_EVENT_ID": event,
        "RFQ_NUMBER": "0010000001",
        "RFQ_ITEM": "00010",
        "QUOTATION_VERSION": version,
        "SUPPLIER_ID": supplier,
        "SUPPLIER_NAME": f"Supplier {supplier[-1]}",
        "MATERIAL_DESCRIPTION": "Synthetic carton",
        "MATERIAL_GROUP": "PACK",
        "PURCHASING_ORG": "1000",
        "REQUESTED_QUANTITY": 1000,
        "QUOTED_QUANTITY": 1000,
        "QUOTATION_UOM": "EA",
        "COMPARISON_UOM": "EA",
        "BASE_UNIT_PRICE": 12.5,
        "PRICE_UNIT": 1,
        "CURRENCY": "INR",
        "QUOTATION_DATE": "2026-07-01",
        "VALIDITY_END_DATE": "2026-12-31",
        "QUOTATION_STATUS": "VALID",
        "SOURCE_TRANSACTION": "ME49",
        "SOURCE_FILE_NAME": "synthetic_rfq.xlsx",
        "SOURCE_EXTRACTED_AT": "2026-07-29T10:00:00",
        "SOURCE_ROW_ID": f"RFQ-{supplier}-{version}",
        "PO_NUMBER": "0045000001",
        "PO_ITEM": "00010",
        "PO_DATE": "2026-06-01",
        "ORDER_QUANTITY": 1000,
        "ORDER_UOM": "EA",
        "NET_PRICE": 12.0,
        "NET_ORDER_VALUE": 12000,
        "PO_STATUS": "COMPLETE",
        "DELETION_FLAG": False,
    }
    return values.get(field, "SYNTHETIC")


def _append_row(
    ws,
    headers: list[str],
    *,
    supplier: str,
    event: str = "EVT-001",
    version: int = 1,
    header_to_canonical: dict[str, str] | None = None,
    overrides: dict[str, object] | None = None,
):
    aliases = header_to_canonical or {}
    updates = overrides or {}
    row = []
    for header in headers:
        canonical = aliases.get(header, header)
        row.append(updates.get(canonical, _value(canonical, supplier, event, version)))
    ws.append(row)


def _metadata_headers() -> list[str]:
    return [
        "UPLOAD_ID",
        "SCHEMA_VERSION",
        "UPLOAD_MODE",
        "SOURCE_SYSTEM",
        "PURCHASING_ORG",
        "BASE_CURRENCY",
        "EXTRACTED_AT",
        "UPLOAD_CREATED_AT",
        "RFQ_SOURCE_TRANSACTION",
        "DATA_CLASSIFICATION",
        "ANONYMIZATION_STATUS",
        "SOURCE_FILE_HASH_SHA256",
        "NOTES",
    ]


def _metadata_values(
    *,
    include_history: bool,
    overrides: dict[str, object] | None = None,
) -> list[object]:
    values = {
        "UPLOAD_ID": "UP-001",
        "SCHEMA_VERSION": "1.3.0",
        "UPLOAD_MODE": "FULL_SOURCING_REVIEW" if include_history else "QUICK_RFQ",
        "SOURCE_SYSTEM": "SAP",
        "PURCHASING_ORG": "1000",
        "BASE_CURRENCY": "INR",
        "EXTRACTED_AT": "2026-07-29T10:00:00",
        "UPLOAD_CREATED_AT": "2026-07-29T10:05:00",
        "RFQ_SOURCE_TRANSACTION": "ME49",
        "DATA_CLASSIFICATION": "SYNTHETIC",
        "ANONYMIZATION_STATUS": "SYNTHETIC",
        "SOURCE_FILE_HASH_SHA256": "a" * 64,
        "NOTES": "Canonical upload hash is calculated externally during ingestion.",
    }
    values.update(overrides or {})
    return [values[header] for header in _metadata_headers()]


def _workbook_bytes(
    *,
    suppliers: tuple[str, ...] = ("0000100001", "0000100002"),
    include_history: bool = True,
    include_metadata: bool = True,
    second_event: bool = False,
    formula_rows: tuple[int, ...] = (),
    row_overrides: dict[int, dict[str, object]] | None = None,
    normalized_alias_header: bool = False,
    exact_alias_header: bool = False,
    metadata_overrides: dict[str, object] | None = None,
    extra_metadata_row: bool = False,
    empty_rfq: bool = False,
) -> bytes:
    wb = Workbook()
    rfq = wb.active
    rfq.title = "RFQ_QUOTES"
    headers = _required_headers("RFQQuoteRow")
    header_to_canonical: dict[str, str] = {}
    if normalized_alias_header:
        index = headers.index("RFQ_NUMBER")
        headers[index] = "RFQ-Number"
        header_to_canonical["RFQ-Number"] = "RFQ_NUMBER"
    if exact_alias_header:
        index = headers.index("SUPPLIER_ID")
        headers[index] = "Vendor Number"
        header_to_canonical["Vendor Number"] = "SUPPLIER_ID"
    rfq.append(headers)

    if not empty_rfq:
        for row_index, supplier in enumerate(suppliers, start=2):
            _append_row(
                rfq,
                headers,
                supplier=supplier,
                header_to_canonical=header_to_canonical,
                overrides=(row_overrides or {}).get(row_index),
            )
        if second_event:
            _append_row(
                rfq,
                headers,
                supplier="0000100003",
                event="EVT-002",
                header_to_canonical=header_to_canonical,
            )
            _append_row(
                rfq,
                headers,
                supplier="0000100004",
                event="EVT-002",
                header_to_canonical=header_to_canonical,
            )
        price_column = headers.index("BASE_UNIT_PRICE") + 1
        for row_number in formula_rows:
            rfq.cell(row=row_number, column=price_column, value="=10+2.5")

    if include_history:
        history = wb.create_sheet("PO_HISTORY")
        history_headers = _required_headers("POHistoryRow")
        history.append(history_headers)
        _append_row(history, history_headers, supplier="0000100001")

    if include_metadata:
        metadata = wb.create_sheet("UPLOAD_METADATA")
        metadata.append(_metadata_headers())
        metadata.append(
            _metadata_values(
                include_history=include_history,
                overrides=metadata_overrides,
            )
        )
        if extra_metadata_row:
            metadata.append(
                _metadata_values(
                    include_history=include_history,
                    overrides={"UPLOAD_ID": "UP-002"},
                )
            )

    stream = BytesIO()
    wb.save(stream)
    wb.close()
    return stream.getvalue()


def _duplicate_workbook(*, contradictory: bool) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "RFQ_QUOTES"
    headers = _required_headers("RFQQuoteRow")
    ws.append(headers)
    _append_row(ws, headers, supplier="0000100001")
    overrides = {
        "SOURCE_ROW_ID": "RFQ-DUPLICATE-2",
        "SOURCE_FILE_NAME": "second_source.xlsx",
        "SOURCE_EXTRACTED_AT": "2026-07-29T11:00:00",
    }
    if contradictory:
        overrides["BASE_UNIT_PRICE"] = 13.5
    _append_row(
        ws,
        headers,
        supplier="0000100001",
        overrides=overrides,
    )
    _append_row(ws, headers, supplier="0000100002")
    stream = BytesIO()
    wb.save(stream)
    wb.close()
    return stream.getvalue()


def test_full_workbook_parses_typed_records_and_hashes():
    payload = _workbook_bytes()
    result = adapt_v13_workbook(payload, filename="PROCUREMENT_COPILOT_UPLOAD.xlsx")

    assert result.mode == "FULL_SOURCING_REVIEW"
    assert result.schema_version == "1.3.0"
    assert result.alias_registry_version == "1.3.0"
    assert result.upload_file_hash_sha256 == sha256(payload).hexdigest()
    assert result.source_file_hash_sha256 == "a" * 64
    assert len(result.rfq_quotes) == 2
    assert len(result.po_history) == 1
    assert result.rfq_quotes[0].canonical_values["RFQ_NUMBER"] == "0010000001"
    assert result.rfq_quotes[0].canonical_values["BASE_UNIT_PRICE"].as_tuple().exponent < 1
    assert result.rfq_quotes[0].normalized_values == {}
    assert result.rfq_quotes[0].valid_for_analysis
    assert result.rfq_quotes[0].provenance.upload_file_hash_sha256 == sha256(payload).hexdigest()
    assert not result.has_fatal


def test_quick_mode_is_detected_without_history():
    result = adapt_v13_workbook(
        _workbook_bytes(include_history=False),
        filename="quick.xlsx",
    )
    assert result.mode == "QUICK_RFQ"
    assert result.po_history == ()
    assert not any(item.code == "PO_HISTORY_UNAVAILABLE" for item in result.findings)


def test_mode_is_inferred_when_metadata_is_absent():
    result = adapt_v13_workbook(
        _workbook_bytes(include_history=False, include_metadata=False),
        filename="quick.xlsx",
    )
    assert result.mode == "QUICK_RFQ"
    assert result.upload_metadata is None


def test_multiple_events_require_explicit_selection():
    payload = _workbook_bytes(second_event=True)
    result = adapt_v13_workbook(payload, filename="multi.xlsx")
    assert result.available_sourcing_event_ids == ("EVT-001", "EVT-002")
    assert any(item.code == "SOURCING_EVENT_SELECTION_REQUIRED" for item in result.findings)

    selected = adapt_v13_workbook(
        payload,
        filename="multi.xlsx",
        selected_sourcing_event_id="EVT-002",
    )
    assert {
        item.canonical_values["SOURCING_EVENT_ID"] for item in selected.rfq_quotes
    } == {"EVT-002"}
    assert not any(
        item.code == "SOURCING_EVENT_SELECTION_REQUIRED"
        for item in selected.findings
    )


def test_invalid_event_selection_is_fatal():
    result = adapt_v13_workbook(
        _workbook_bytes(),
        filename="event.xlsx",
        selected_sourcing_event_id="MISSING",
    )
    assert any(
        item.code == "SOURCING_EVENT_SELECTION_INVALID"
        and item.severity == "Fatal"
        for item in result.findings
    )


def test_formula_in_one_price_blocks_row_and_supplier_gate():
    result = adapt_v13_workbook(
        _workbook_bytes(formula_rows=(2,)),
        filename="formula.xlsx",
    )
    assert any(
        item.code == "FORMULA_CELL_REJECTED"
        and item.field_name == "BASE_UNIT_PRICE"
        for item in result.findings
    )
    assert not result.rfq_quotes[0].valid_for_analysis
    assert result.rfq_quotes[0].canonical_values.get("BASE_UNIT_PRICE") is None
    assert any(
        item.code == "MINIMUM_SUPPLIER_COUNT_NOT_MET"
        for item in result.findings
    )
    assert not any(
        item.code == "NO_VALID_QUOTATION_RECORDS"
        for item in result.findings
    )


def test_all_formula_prices_are_fatal_no_valid_records():
    result = adapt_v13_workbook(
        _workbook_bytes(formula_rows=(2, 3)),
        filename="all-formulas.xlsx",
    )
    assert all(not item.valid_for_analysis for item in result.rfq_quotes)
    assert any(
        item.code == "NO_VALID_QUOTATION_RECORDS"
        and item.severity == "Fatal"
        for item in result.findings
    )


def test_all_invalid_values_are_fatal_no_valid_records():
    result = adapt_v13_workbook(
        _workbook_bytes(
            row_overrides={
                2: {"REQUESTED_QUANTITY": 0},
                3: {"REQUESTED_QUANTITY": 0},
            }
        ),
        filename="all-invalid.xlsx",
    )
    assert any(
        item.code == "NO_VALID_QUOTATION_RECORDS"
        and item.severity == "Fatal"
        for item in result.findings
    )


def test_three_suppliers_with_one_blocked_still_meet_supplier_gate():
    result = adapt_v13_workbook(
        _workbook_bytes(
            suppliers=("0000100001", "0000100002", "0000100003"),
            formula_rows=(2,),
        ),
        filename="three-suppliers.xlsx",
    )
    assert sum(item.valid_for_analysis for item in result.rfq_quotes) == 2
    assert not any(
        item.code == "MINIMUM_SUPPLIER_COUNT_NOT_MET"
        for item in result.findings
    )


def test_superseded_version_does_not_count_as_an_extra_supplier():
    wb = Workbook()
    ws = wb.active
    ws.title = "RFQ_QUOTES"
    headers = _required_headers("RFQQuoteRow")
    ws.append(headers)
    _append_row(ws, headers, supplier="0000100001", version=1)
    _append_row(ws, headers, supplier="0000100001", version=2)
    stream = BytesIO()
    wb.save(stream)
    wb.close()

    result = adapt_v13_workbook(stream.getvalue(), filename="versions.xlsx")
    assert sum(item.active for item in result.rfq_quotes) == 1
    assert any(
        item.code == "MINIMUM_SUPPLIER_COUNT_NOT_MET"
        for item in result.findings
    )


def test_normalized_high_risk_alias_requires_confirmation_and_blocks_rows():
    payload = _workbook_bytes(normalized_alias_header=True)
    result = adapt_v13_workbook(payload, filename="alias.xlsx")
    review = next(
        item for item in result.mapping_reviews if item.source_header == "RFQ-Number"
    )
    assert review.canonical_field == "RFQ_NUMBER"
    assert review.confidence_class == "NORMALIZED_APPROVED"
    assert review.requires_confirmation
    assert all(not item.valid_for_analysis for item in result.rfq_quotes)
    assert any(
        item.code == "HIGH_RISK_MAPPING_CONFIRMATION_REQUIRED"
        for item in result.findings
    )

    confirmed = adapt_v13_workbook(
        payload,
        filename="alias.xlsx",
        confirmed_mappings={("RFQ_QUOTES", "RFQ-Number", "RFQ_NUMBER")},
    )
    confirmed_review = next(
        item
        for item in confirmed.mapping_reviews
        if item.source_header == "RFQ-Number"
    )
    assert not confirmed_review.requires_confirmation
    assert all(item.valid_for_analysis for item in confirmed.rfq_quotes)


def test_exact_sap_alias_maps_without_confirmation():
    result = adapt_v13_workbook(
        _workbook_bytes(exact_alias_header=True),
        filename="exact-alias.xlsx",
    )
    review = next(
        item
        for item in result.mapping_reviews
        if item.source_header == "Vendor Number"
    )
    assert review.canonical_field == "SUPPLIER_ID"
    assert review.confidence_class == "EXACT_APPROVED"
    assert not review.requires_confirmation
    assert {
        item.canonical_values["SUPPLIER_ID"] for item in result.rfq_quotes
    } == {"0000100001", "0000100002"}


def test_latest_quotation_version_is_active_and_prior_version_retained():
    wb = Workbook()
    ws = wb.active
    ws.title = "RFQ_QUOTES"
    headers = _required_headers("RFQQuoteRow")
    ws.append(headers)
    _append_row(ws, headers, supplier="0000100001", version=1)
    _append_row(ws, headers, supplier="0000100001", version=2)
    _append_row(ws, headers, supplier="0000100002", version=1)
    stream = BytesIO()
    wb.save(stream)
    wb.close()

    result = adapt_v13_workbook(stream.getvalue(), filename="versions.xlsx")
    supplier_one = [
        item
        for item in result.rfq_quotes
        if item.canonical_values["SUPPLIER_ID"] == "0000100001"
    ]
    assert len(supplier_one) == 2
    assert [
        item.canonical_values["QUOTATION_VERSION"]
        for item in supplier_one
        if item.active
    ] == [2]


def test_provenance_only_duplicate_is_warning():
    result = adapt_v13_workbook(
        _duplicate_workbook(contradictory=False),
        filename="duplicate.xlsx",
    )
    assert any(
        item.code == "EXACT_DUPLICATE_ROW"
        and item.severity == "Warning"
        for item in result.findings
    )
    assert not any(
        item.code == "CONTRADICTORY_DUPLICATE_KEY"
        for item in result.findings
    )


def test_contradictory_business_duplicate_is_fatal():
    result = adapt_v13_workbook(
        _duplicate_workbook(contradictory=True),
        filename="contradictory.xlsx",
    )
    assert any(
        item.code == "CONTRADICTORY_DUPLICATE_KEY"
        and item.severity == "Fatal"
        for item in result.findings
    )


def test_empty_rfq_data_section_is_fatal():
    result = adapt_v13_workbook(
        _workbook_bytes(empty_rfq=True),
        filename="empty-rfq.xlsx",
    )
    assert any(
        item.code == "RFQ_QUOTES_EMPTY"
        and item.severity == "Fatal"
        for item in result.findings
    )


def test_multiple_metadata_rows_are_fatal():
    result = adapt_v13_workbook(
        _workbook_bytes(extra_metadata_row=True),
        filename="metadata-cardinality.xlsx",
    )
    assert any(
        item.code == "UPLOAD_METADATA_CARDINALITY_INVALID"
        and item.severity == "Fatal"
        for item in result.findings
    )


def test_malformed_source_hash_is_blocking():
    result = adapt_v13_workbook(
        _workbook_bytes(
            metadata_overrides={"SOURCE_FILE_HASH_SHA256": "not-a-hash"}
        ),
        filename="bad-hash.xlsx",
    )
    assert any(
        item.code == "SOURCE_FILE_HASH_INVALID"
        and item.severity == "Blocking"
        for item in result.findings
    )


def test_unsupported_schema_version_is_fatal():
    result = adapt_v13_workbook(
        _workbook_bytes(metadata_overrides={"SCHEMA_VERSION": "1.2.0"}),
        filename="bad-schema.xlsx",
    )
    assert any(
        item.code == "SCHEMA_VERSION_UNSUPPORTED"
        and item.severity == "Fatal"
        for item in result.findings
    )


def test_invalid_anonymization_and_currency_are_blocking():
    result = adapt_v13_workbook(
        _workbook_bytes(
            metadata_overrides={
                "ANONYMIZATION_STATUS": "UNKNOWN",
                "BASE_CURRENCY": "inr",
            }
        ),
        filename="metadata-values.xlsx",
    )
    assert any(
        item.code == "ANONYMIZATION_STATUS_INVALID"
        for item in result.findings
    )
    assert any(
        item.code == "CURRENCY_FORMAT_INVALID"
        and item.sheet == "UPLOAD_METADATA"
        for item in result.findings
    )


def test_selected_event_with_all_rows_blocked_is_fatal():
    payload = _workbook_bytes(
        second_event=True,
        formula_rows=(4, 5),
    )
    result = adapt_v13_workbook(
        payload,
        filename="selected-invalid.xlsx",
        selected_sourcing_event_id="EVT-002",
    )
    assert any(
        item.code == "NO_VALID_SELECTED_EVENT_QUOTATIONS"
        and item.severity == "Fatal"
        for item in result.findings
    )


def test_renamed_non_xlsx_payload_is_rejected_by_existing_safety_loader():
    with pytest.raises(WorkbookAdapterError, match="valid XLSX package"):
        adapt_v13_workbook(b"not an xlsx", filename="renamed.xlsx")


def test_alias_contract_is_versioned_and_collective_number_is_not_event_alias():
    registry = json.loads(Path(ALIAS_PATH).read_text(encoding="utf-8"))
    assert registry["registry_version"] == "1.3.0"
    assert (
        "Collective Number"
        not in registry["sheets"]["RFQ_QUOTES"]["SOURCING_EVENT_ID"]
    )
    assert (
        "Collective Number"
        in registry["sheets"]["RFQ_QUOTES"]["COLLECTIVE_NUMBER"]
    )
