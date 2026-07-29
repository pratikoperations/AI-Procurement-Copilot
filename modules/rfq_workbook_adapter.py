"""Governed adapter for the versioned v1.3 procurement workbook."""
from __future__ import annotations

from dataclasses import dataclass, replace
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from io import BytesIO
import json
from pathlib import Path
import re
from typing import Any, BinaryIO, Iterable, Mapping, Sequence

from openpyxl import load_workbook

from modules.erp_workbook_loader import (
    DEFAULT_MAX_FILE_SIZE_BYTES,
    WorkbookLoadError,
    load_erp_workbook,
)
from modules.ranking_input_contract import RankingContractError, load_contract_bundle
from modules.ranking_input_matching import calculate_mode_eligibility, cross_row_findings, match_ranking_records
from modules.ranking_input_models import (
    CanonicalFieldEvidenceResult,
    CanonicalRankingRecord,
    RankingMappingConfirmation,
    RankingModeEligibility,
    RankingScopeMatch,
    VALUE_ORIGINS,
)
from modules.ranking_input_semantics import generate_evidence_results

CONTRACT_ROOT = Path(__file__).resolve().parents[1] / "planning" / "v1.3" / "build_group_b"
SCHEMA_PATH = CONTRACT_ROOT / "minimum_workbook_schema_v1.3.0.json"
ALIAS_PATH = CONTRACT_ROOT / "sap_report_alias_registry_v1.3.0.json"
APPROVED_SHEETS = ("RFQ_QUOTES", "PO_HISTORY", "UPLOAD_METADATA")
ROW_DEFS = {"RFQ_QUOTES": "RFQQuoteRow", "PO_HISTORY": "POHistoryRow", "UPLOAD_METADATA": "UploadMetadataRow"}
HIGH_RISK = {
    "SOURCING_EVENT_ID", "RFQ_NUMBER", "RFQ_ITEM", "SUPPLIER_ID", "QUOTATION_VERSION",
    "REQUESTED_QUANTITY", "QUOTED_QUANTITY", "ORDER_QUANTITY", "BASE_UNIT_PRICE",
    "NET_PRICE", "PRICE_UNIT", "CURRENCY", "QUOTATION_UOM", "ORDER_UOM",
    "COMPARISON_UOM", "UOM_CONVERSION_FACTOR", "EXCHANGE_RATE", "EXCHANGE_RATE_DATE",
}
FORMULA_BLOCKING = HIGH_RISK | {
    "QUOTATION_DATE", "VALIDITY_END_DATE", "QUOTATION_STATUS", "PO_DATE", "PO_STATUS",
    "PO_NUMBER", "PO_ITEM", "SOURCE_ROW_ID",
}
PROVENANCE_ONLY_FIELDS = {"SOURCE_ROW_ID", "SOURCE_FILE_NAME", "SOURCE_EXTRACTED_AT"}
SHEET_LEVEL_ROW_BLOCKING_CODES = {
    "AMBIGUOUS_HEADER_MAPPING", "DUPLICATE_CANONICAL_TARGET",
    "HIGH_RISK_MAPPING_CONFIRMATION_REQUIRED", "MANDATORY_HEADER_MISSING",
}
VALIDITY_SEVERITIES = {"Fatal", "Blocking"}
RANKING_SHEET = "SUPPLIER_RANKING_INPUTS"


@dataclass(frozen=True)
class Finding:
    severity: str
    code: str
    message: str
    sheet: str | None = None
    row_number: int | None = None
    field_name: str | None = None


@dataclass(frozen=True)
class MappingReview:
    sheet: str
    source_header: str
    canonical_field: str | None
    confidence_class: str
    source_classification: str | None
    requires_confirmation: bool
    reason: str | None = None


@dataclass(frozen=True)
class Provenance:
    source_filename: str
    sheet: str
    source_row_number: int
    source_row_id: str
    source_file_hash_sha256: str | None
    upload_file_hash_sha256: str
    schema_version: str
    alias_registry_version: str


@dataclass(frozen=True)
class CanonicalRecord:
    sheet: str
    original_values: Mapping[str, Any]
    canonical_values: Mapping[str, Any]
    normalized_values: Mapping[str, Any]
    provenance: Provenance
    active: bool = False
    row_valid: bool = True
    eligible_for_analysis: bool = True

    @property
    def valid_for_analysis(self) -> bool:
        return self.eligible_for_analysis


@dataclass(frozen=True)
class AdapterResult:
    filename: str
    mode: str
    schema_version: str
    alias_registry_version: str
    upload_file_hash_sha256: str
    source_file_hash_sha256: str | None
    selected_sourcing_event_id: str | None
    available_sourcing_event_ids: tuple[str, ...]
    rfq_quotes: tuple[CanonicalRecord, ...]
    po_history: tuple[CanonicalRecord, ...]
    upload_metadata: Mapping[str, Any] | None
    mapping_reviews: tuple[MappingReview, ...]
    findings: tuple[Finding, ...]
    supplier_ranking_inputs: tuple[CanonicalRankingRecord, ...] = ()
    ranking_evidence_results: tuple[CanonicalFieldEvidenceResult, ...] = ()
    ranking_scope_matches: tuple[RankingScopeMatch, ...] = ()
    ranking_mode_eligibility: tuple[RankingModeEligibility, ...] = ()

    @property
    def has_fatal(self) -> bool:
        return any(item.severity == "Fatal" for item in self.findings)

    @property
    def has_blocking(self) -> bool:
        return any(item.severity == "Blocking" for item in self.findings)


class WorkbookAdapterError(ValueError):
    """Raised when the workbook cannot enter the governed adapter."""


def _read_bytes(source: str | Path | bytes | bytearray | BinaryIO) -> bytes:
    if isinstance(source, (str, Path)):
        return Path(source).read_bytes()
    if isinstance(source, (bytes, bytearray)):
        return bytes(source)
    read = getattr(source, "read", None)
    if not callable(read):
        raise WorkbookAdapterError("Workbook source must be a path, bytes, or binary file object.")
    position = source.tell() if callable(getattr(source, "tell", None)) else None
    try:
        if callable(getattr(source, "seek", None)):
            source.seek(0)
        payload = read()
    finally:
        if position is not None and callable(getattr(source, "seek", None)):
            source.seek(position)
    if not isinstance(payload, (bytes, bytearray)):
        raise WorkbookAdapterError("Workbook file object must return bytes.")
    return bytes(payload)


def _filename(source: object, supplied: str | None) -> str:
    if supplied:
        return Path(supplied).name
    if isinstance(source, (str, Path)):
        return Path(source).name
    return Path(str(getattr(source, "name", "PROCUREMENT_COPILOT_UPLOAD.xlsx"))).name


def _load_contract(path: Path) -> dict[str, Any]:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise WorkbookAdapterError(f"Contract file could not be loaded: {path}") from exc


def _normalise_header(value: Any) -> str:
    return re.sub(r"[\W_\s]+", " ", "" if value is None else str(value).strip()).casefold().strip()


def _canonical_alias_index(sheet: str, schema: Mapping[str, Any], registry: Mapping[str, Any]) -> dict[str, list[str]]:
    properties = schema["$defs"][ROW_DEFS[sheet]].get("properties", {})
    aliases = registry.get("sheets", {}).get(sheet, {})
    index: dict[str, list[str]] = {}
    for canonical in properties:
        for label in (canonical, *aliases.get(canonical, [])):
            index.setdefault(_normalise_header(label), []).append(canonical)
    return index


def _map_headers(sheet: str, headers: Sequence[Any], schema: Mapping[str, Any], registry: Mapping[str, Any], confirmed: set[tuple[str, str, str]], findings: list[Finding]) -> tuple[dict[int, str], list[MappingReview]]:
    index = _canonical_alias_index(sheet, schema, registry)
    aliases = registry.get("sheets", {}).get(sheet, {})
    classes = registry.get("field_source_classifications", {}).get(sheet, {})
    mapped: dict[int, str] = {}
    reviews: list[MappingReview] = []
    used_targets: set[str] = set()
    for position, raw in enumerate(headers):
        source = "" if raw is None else str(raw).strip()
        candidates = tuple(dict.fromkeys(index.get(_normalise_header(raw), [])))
        if not source:
            reviews.append(MappingReview(sheet, source, None, "UNMAPPED", None, False, "Blank header")); continue
        if len(candidates) > 1:
            reviews.append(MappingReview(sheet, source, None, "AMBIGUOUS", None, True, f"Matches {candidates}"))
            findings.append(Finding("Fatal", "AMBIGUOUS_HEADER_MAPPING", f"Header '{source}' maps to multiple fields.", sheet)); continue
        if not candidates:
            reviews.append(MappingReview(sheet, source, None, "UNMAPPED", None, False, "No approved alias"))
            findings.append(Finding("Information", "UNMAPPED_SOURCE_COLUMN", f"Column '{source}' is ignored by calculations.", sheet)); continue
        canonical = candidates[0]
        if canonical in used_targets:
            reviews.append(MappingReview(sheet, source, canonical, "AMBIGUOUS", classes.get(canonical), True, "Duplicate canonical target"))
            findings.append(Finding("Fatal", "DUPLICATE_CANONICAL_TARGET", f"Multiple columns map to '{canonical}'.", sheet, field_name=canonical)); continue
        approved = {str(item).strip().casefold() for item in (canonical, *aliases.get(canonical, []))}
        confidence = "EXACT_APPROVED" if source.casefold() in approved else "NORMALIZED_APPROVED"
        needs_confirmation = canonical in HIGH_RISK and confidence != "EXACT_APPROVED"
        if (sheet, source, canonical) in confirmed:
            needs_confirmation = False
        reviews.append(MappingReview(sheet, source, canonical, confidence, classes.get(canonical), needs_confirmation, "High-risk normalized mapping" if needs_confirmation else None))
        if needs_confirmation:
            findings.append(Finding("Blocking", "HIGH_RISK_MAPPING_CONFIRMATION_REQUIRED", f"Mapping '{source}' to '{canonical}' requires confirmation.", sheet, field_name=canonical))
        mapped[position] = canonical; used_targets.add(canonical)
    return mapped, reviews


def _schema_type(property_schema: Mapping[str, Any]) -> str | None:
    reference = property_schema.get("$ref")
    if reference:
        return reference.rsplit("/", 1)[-1]
    kind = property_schema.get("type")
    if isinstance(kind, list):
        return next((item for item in kind if item != "null"), None)
    for option in property_schema.get("anyOf", []):
        if "$ref" in option:
            return option["$ref"].rsplit("/", 1)[-1]
        if option.get("type") != "null":
            return option.get("type")
    return kind


def _coerce(value: Any, property_schema: Mapping[str, Any]) -> Any:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    kind = _schema_type(property_schema)
    if kind in {"string", "nonEmptyString", "evidenceStatus", "valueOrigin"}:
        return str(value).strip()
    if kind == "integer":
        number = Decimal(str(value))
        if isinstance(value, bool) or number != number.to_integral_value():
            raise ValueError("must be an integer")
        return int(number)
    if kind in {"number", "percent", "score"}:
        if isinstance(value, bool):
            raise ValueError("must be numeric")
        return Decimal(str(value))
    if kind == "boolean":
        if isinstance(value, bool):
            return value
        text = str(value).strip().casefold()
        if text in {"true", "yes", "y", "1", "x"}: return True
        if text in {"false", "no", "n", "0"}: return False
        raise ValueError("must be a recognised boolean")
    if kind == "date":
        if isinstance(value, datetime): return value.date()
        if isinstance(value, date): return value
        return date.fromisoformat(str(value).strip())
    if kind == "datetime":
        if isinstance(value, datetime): return value
        if isinstance(value, date): return datetime.combine(value, datetime.min.time())
        return datetime.fromisoformat(str(value).strip().replace("Z", "+00:00"))
    if kind == "object":
        if isinstance(value, Mapping): return dict(value)
        parsed = json.loads(str(value))
        if not isinstance(parsed, Mapping): raise ValueError("must be a JSON object")
        return dict(parsed)
    return value


def _empty(values: Iterable[Any]) -> bool:
    return all(value is None or (isinstance(value, str) and not value.strip()) for value in values)


def _parse_sheet(workbook: Any, sheet: str, filename: str, upload_hash: str, source_hash: str | None, schema: Mapping[str, Any], registry: Mapping[str, Any], confirmed: set[tuple[str, str, str]], findings: list[Finding]) -> tuple[list[CanonicalRecord], list[MappingReview]]:
    if sheet not in workbook.sheetnames:
        return [], []
    rows = workbook[sheet].iter_rows()
    header_cells = next(rows, None)
    if not header_cells:
        findings.append(Finding("Fatal", "EMPTY_SHEET", f"Sheet '{sheet}' has no header row.", sheet)); return [], []
    headers = [cell.value for cell in header_cells]
    mapped, reviews = _map_headers(sheet, headers, schema, registry, confirmed, findings)
    row_schema = schema["$defs"][ROW_DEFS[sheet]]
    required = set(row_schema.get("required", []))
    for missing in sorted(required - set(mapped.values())):
        findings.append(Finding("Fatal", "MANDATORY_HEADER_MISSING", f"Mandatory field '{missing}' is not mapped.", sheet, field_name=missing))
    records: list[CanonicalRecord] = []
    for row_number, cells in enumerate(rows, start=2):
        values = [cell.value for cell in cells]
        if _empty(values): continue
        formulas = {mapped[index] for index, cell in enumerate(cells) if index in mapped and cell.data_type == "f"}
        for field_name in formulas:
            findings.append(Finding("Blocking" if field_name in FORMULA_BLOCKING else "Warning", "FORMULA_CELL_REJECTED", f"Formula is not accepted for '{field_name}'.", sheet, row_number, field_name))
        original: dict[str, Any] = {}; canonical: dict[str, Any] = {}
        for index, value in enumerate(values):
            source_header = str(headers[index]) if index < len(headers) else f"COLUMN_{index + 1}"
            original[source_header] = value
            canonical_name = mapped.get(index)
            if canonical_name is None or canonical_name in formulas: continue
            try:
                canonical[canonical_name] = _coerce(value, row_schema.get("properties", {}).get(canonical_name, {}))
            except (ValueError, TypeError, InvalidOperation, json.JSONDecodeError) as exc:
                severity = "Fatal" if canonical_name in required else "Blocking"
                findings.append(Finding(severity, "VALUE_TYPE_INVALID", f"Invalid '{canonical_name}': {exc}", sheet, row_number, canonical_name)); canonical[canonical_name] = None
        for required_field in required:
            if canonical.get(required_field) is None:
                findings.append(Finding("Fatal", "MANDATORY_VALUE_MISSING", f"Mandatory value '{required_field}' is missing.", sheet, row_number, required_field))
        row_id = str(canonical.get("SOURCE_ROW_ID") or f"{sheet}:{row_number}")
        records.append(CanonicalRecord(sheet, original, canonical, {}, Provenance(filename, sheet, row_number, row_id, source_hash, upload_hash, str(schema.get("version", "unknown")), str(registry.get("registry_version", "unknown")))))
    return records, reviews


def _detect_schema_version(workbook: Any) -> str:
    if "UPLOAD_METADATA" not in workbook.sheetnames: return "1.3.0"
    rows = workbook["UPLOAD_METADATA"].iter_rows(values_only=True)
    headers = next(rows, None); first = next(rows, None)
    if not headers or not first: return "1.3.0"
    index = {str(value).strip(): position for position, value in enumerate(headers) if value is not None}
    position = index.get("SCHEMA_VERSION")
    if position is None or position >= len(first) or first[position] in (None, ""): return "1.3.0"
    return str(first[position]).strip()


def _ranking_alias_index(row_schema: Mapping[str, Any], registry: Mapping[str, Any]) -> dict[str, list[str]]:
    aliases = registry.get("sheets", {}).get(RANKING_SHEET, {})
    index: dict[str, list[str]] = {}
    for canonical in row_schema.get("properties", {}):
        for label in (canonical, *aliases.get(canonical, [])):
            index.setdefault(_normalise_header(label), []).append(canonical)
    return index


def _confirmation_valid(confirmations: Sequence[RankingMappingConfirmation], *, upload_hash: str, source_header: str, canonical_field: str, actual_origin: str | None) -> bool:
    return any(
        item.upload_hash_sha256 == upload_hash and item.schema_version == "1.3.1"
        and item.alias_registry_version == "1.3.1" and item.sheet == RANKING_SHEET
        and item.source_header == source_header and item.canonical_field == canonical_field
        and item.detected_scale == "0_TO_100_ONLY" and item.value_origin == actual_origin
        for item in confirmations
    )


def _parse_ranking_sheet(workbook: Any, filename: str, upload_hash: str, source_hash: str | None, schema: Mapping[str, Any], registry: Mapping[str, Any], confirmations: Sequence[RankingMappingConfirmation], findings: list[Finding]) -> tuple[list[CanonicalRankingRecord], list[MappingReview], set[tuple[str, str]], set[str]]:
    if RANKING_SHEET not in workbook.sheetnames:
        findings.append(Finding("Blocking", "SUPPLIER_RANKING_INPUTS_SHEET_MISSING", "v1.3.1 workbook has no supplier ranking-input sheet.", RANKING_SHEET)); return [], [], set(), set()
    row_schema = schema["$defs"]["SupplierRankingInputRow"]
    rejected = {_normalise_header(key): key for key in registry.get("rejected_semantic_aliases", {})}
    index = _ranking_alias_index(row_schema, registry)
    rows = workbook[RANKING_SHEET].iter_rows(); header_cells = next(rows, None)
    if not header_cells:
        findings.append(Finding("Fatal", "RANKING_REQUIRED_HEADER_MISSING", "Ranking sheet has no header row.", RANKING_SHEET)); return [], [], set(), set()
    headers = ["" if cell.value is None else str(cell.value).strip() for cell in header_cells]
    mapped: dict[int, str] = {}; alias_sources: dict[str, str] = {}; used: set[str] = set(); preliminary: list[MappingReview] = []
    for position, source in enumerate(headers):
        normalized = _normalise_header(source)
        if normalized in rejected:
            preliminary.append(MappingReview(RANKING_SHEET, source, None, "REJECTED", None, True, "Rejected semantic alias")); findings.append(Finding("Blocking", "RANKING_FIELD_SEMANTICS_UNCONFIRMED", f"Header '{source}' is a rejected semantic alias.", RANKING_SHEET)); continue
        candidates = tuple(dict.fromkeys(index.get(normalized, [])))
        if len(candidates) > 1:
            preliminary.append(MappingReview(RANKING_SHEET, source, None, "AMBIGUOUS", None, True, "Multiple candidates")); findings.append(Finding("Fatal", "DUPLICATE_RANKING_CANONICAL_TARGET", f"Header '{source}' maps ambiguously.", RANKING_SHEET)); continue
        if not candidates:
            preliminary.append(MappingReview(RANKING_SHEET, source, None, "UNMAPPED", None, False, "Exact canonical header required")); findings.append(Finding("Information", "UNLISTED_RANKING_HEADER_IGNORED", f"Unlisted ranking header '{source}' is ignored.", RANKING_SHEET)); continue
        canonical = candidates[0]
        if canonical in used:
            preliminary.append(MappingReview(RANKING_SHEET, source, canonical, "AMBIGUOUS", None, True, "Duplicate canonical target")); findings.append(Finding("Fatal", "DUPLICATE_RANKING_CANONICAL_TARGET", f"Multiple columns map to '{canonical}'.", RANKING_SHEET, field_name=canonical)); continue
        mapped[position] = canonical; used.add(canonical)
        if source != canonical: alias_sources[canonical] = source
    required = set(row_schema.get("required", []))
    for missing in sorted(required - set(mapped.values())):
        findings.append(Finding("Fatal", "RANKING_REQUIRED_HEADER_MISSING", f"Mandatory ranking header '{missing}' is not mapped.", RANKING_SHEET, field_name=missing))
    records: list[CanonicalRankingRecord] = []
    for row_number, cells in enumerate(rows, start=2):
        raw_values = [cell.value for cell in cells]
        if _empty(raw_values): continue
        canonical: dict[str, Any] = {}
        for position, value in enumerate(raw_values):
            field = mapped.get(position)
            if field is None: continue
            try: canonical[field] = _coerce(value, row_schema.get("properties", {}).get(field, {}))
            except (ValueError, TypeError, InvalidOperation, json.JSONDecodeError) as exc:
                canonical[field] = None; findings.append(Finding("Blocking", "RANKING_INPUT_INVALID_TYPE", f"Invalid '{field}': {exc}", RANKING_SHEET, row_number, field))
        for field in required:
            if canonical.get(field) is None:
                findings.append(Finding("Fatal", "RANKING_REQUIRED_HEADER_MISSING", f"Mandatory ranking value '{field}' is missing.", RANKING_SHEET, row_number, field))
        row_id = str(canonical.get("SOURCE_ROW_ID") or canonical.get("RANKING_INPUT_RECORD_ID") or f"{RANKING_SHEET}:{row_number}")
        origins = canonical.get("VALUE_ORIGINS") if isinstance(canonical.get("VALUE_ORIGINS"), Mapping) else {}
        provenance = Provenance(filename, RANKING_SHEET, row_number, row_id, source_hash, upload_hash, "1.3.1", "1.3.1")
        records.append(CanonicalRankingRecord(canonical, dict(origins), canonical.get("SOURCE_EVIDENCE_STATUS"), provenance, True, bool(canonical.get("ACTIVE_FLAG", True))))
    confirmed_origins: set[tuple[str, str]] = set(); pending_suppliers: set[str] = set(); reviews = list(preliminary)
    for canonical, source in alias_sources.items():
        missing_context = False
        for record in records:
            if record.canonical_values.get(canonical) is None: continue
            origin = record.value_origins.get(canonical)
            if _confirmation_valid(confirmations, upload_hash=upload_hash, source_header=source, canonical_field=canonical, actual_origin=origin):
                if origin: confirmed_origins.add((canonical, origin))
            else:
                missing_context = True; pending_suppliers.add(str(record.canonical_values.get("SUPPLIER_ID") or ""))
                findings.append(Finding("Blocking", "RANKING_MAPPING_CONFIRMATION_REQUIRED", f"Mapping '{source}' to '{canonical}' requires confirmation for origin '{origin}'.", RANKING_SHEET, record.provenance.source_row_number, canonical))
        reviews.append(MappingReview(RANKING_SHEET, source, canonical, "APPROVED_ALIAS", None, missing_context, "Every alias/origin context requires confirmation" if missing_context else None))
    for canonical in set(row_schema.get("properties", {})) - set(alias_sources):
        if canonical in used: reviews.append(MappingReview(RANKING_SHEET, canonical, canonical, "EXACT_CANONICAL", None, False, None))
    updated = [replace(record, row_valid=not any(item.severity in VALIDITY_SEVERITIES and item.sheet == RANKING_SHEET and item.row_number == record.provenance.source_row_number for item in findings)) for record in records]
    return updated, reviews, confirmed_origins, pending_suppliers


def _validate_values(records: Sequence[CanonicalRecord], findings: list[Finding]) -> None:
    positive = {"REQUESTED_QUANTITY", "QUOTED_QUANTITY", "ORDER_QUANTITY", "PRICE_UNIT", "UOM_CONVERSION_FACTOR", "EXCHANGE_RATE"}
    non_negative = {"BASE_UNIT_PRICE", "NET_PRICE", "NET_ORDER_VALUE", "FREIGHT_AMOUNT", "PACKING_AMOUNT", "INSURANCE_AMOUNT", "DUTY_AMOUNT", "TAX_AMOUNT", "TOOLING_AMOUNT", "OTHER_CHARGES_AMOUNT"}
    for record in records:
        for field_name in positive:
            value = record.canonical_values.get(field_name)
            if value is not None and value <= 0: findings.append(Finding("Blocking", "POSITIVE_VALUE_REQUIRED", f"'{field_name}' must be greater than zero.", record.sheet, record.provenance.source_row_number, field_name))
        for field_name in non_negative:
            value = record.canonical_values.get(field_name)
            if value is not None and value < 0: findings.append(Finding("Blocking", "NEGATIVE_VALUE_REJECTED", f"'{field_name}' cannot be negative.", record.sheet, record.provenance.source_row_number, field_name))


def _business_payload(record: CanonicalRecord) -> dict[str, Any]:
    return {key: value for key, value in record.canonical_values.items() if key not in PROVENANCE_ONLY_FIELDS}


def _validate_keys(records: Sequence[CanonicalRecord], findings: list[Finding]) -> None:
    key_fields = {"RFQ_QUOTES": ("SOURCING_EVENT_ID", "RFQ_NUMBER", "RFQ_ITEM", "SUPPLIER_ID", "QUOTATION_VERSION"), "PO_HISTORY": ("PO_NUMBER", "PO_ITEM")}
    seen: dict[tuple[str, tuple[Any, ...]], CanonicalRecord] = {}
    for record in records:
        fields = key_fields.get(record.sheet)
        if not fields: continue
        key = tuple(record.canonical_values.get(name) for name in fields); compound = (record.sheet, key); prior = seen.get(compound)
        if prior is None: seen[compound] = record
        elif _business_payload(prior) != _business_payload(record): findings.append(Finding("Fatal", "CONTRADICTORY_DUPLICATE_KEY", f"Contradictory duplicate key {key}.", record.sheet, record.provenance.source_row_number))
        else: findings.append(Finding("Warning", "EXACT_DUPLICATE_ROW", f"Duplicate business payload for key {key}.", record.sheet, record.provenance.source_row_number))


def _validate_metadata(records: Sequence[CanonicalRecord], schema: Mapping[str, Any], findings: list[Finding]) -> Mapping[str, Any] | None:
    if len(records) > 1: findings.append(Finding("Fatal", "UPLOAD_METADATA_CARDINALITY_INVALID", "UPLOAD_METADATA may contain at most one non-empty data row.", "UPLOAD_METADATA"))
    if not records: return None
    metadata = dict(records[0].canonical_values); constraints = schema.get("x-field-constraints", {}); row = records[0].provenance.source_row_number
    expected = str(constraints.get("schema_version_const", schema.get("version", "1.3.0")))
    if str(metadata.get("SCHEMA_VERSION") or "") != expected: findings.append(Finding("Fatal", "SCHEMA_VERSION_UNSUPPORTED", f"SCHEMA_VERSION must equal '{expected}'.", "UPLOAD_METADATA", row, "SCHEMA_VERSION"))
    modes = set(constraints.get("upload_mode_enum", ())); mode = metadata.get("UPLOAD_MODE")
    if mode is not None and modes and mode not in modes: findings.append(Finding("Fatal", "UPLOAD_MODE_INVALID", f"Unsupported upload mode '{mode}'.", "UPLOAD_METADATA", row, "UPLOAD_MODE"))
    anonymous = set(constraints.get("anonymization_enum", ())); status = metadata.get("ANONYMIZATION_STATUS")
    if status is not None and anonymous and status not in anonymous: findings.append(Finding("Blocking", "ANONYMIZATION_STATUS_INVALID", f"ANONYMIZATION_STATUS must be one of {sorted(anonymous)}.", "UPLOAD_METADATA", row, "ANONYMIZATION_STATUS"))
    source_hash = metadata.get("SOURCE_FILE_HASH_SHA256")
    if source_hash is None or re.fullmatch(constraints.get("sha256_pattern", r"^[a-fA-F0-9]{64}$"), str(source_hash)) is None: findings.append(Finding("Blocking", "SOURCE_FILE_HASH_INVALID", "SOURCE_FILE_HASH_SHA256 must contain exactly 64 hexadecimal characters.", "UPLOAD_METADATA", row, "SOURCE_FILE_HASH_SHA256"))
    currency = metadata.get("BASE_CURRENCY")
    if currency is not None and re.fullmatch(constraints.get("currency_pattern", r"^[A-Z]{3}$"), str(currency)) is None: findings.append(Finding("Blocking", "CURRENCY_FORMAT_INVALID", "BASE_CURRENCY must be a three-letter uppercase currency code.", "UPLOAD_METADATA", row, "BASE_CURRENCY"))
    return metadata


def _validate_row_currencies(records: Sequence[CanonicalRecord], schema: Mapping[str, Any], findings: list[Finding]) -> None:
    pattern = schema.get("x-field-constraints", {}).get("currency_pattern", r"^[A-Z]{3}$")
    for record in records:
        currency = record.canonical_values.get("CURRENCY")
        if currency is not None and re.fullmatch(pattern, str(currency)) is None: findings.append(Finding("Blocking", "CURRENCY_FORMAT_INVALID", "CURRENCY must be a three-letter uppercase currency code.", record.sheet, record.provenance.source_row_number, "CURRENCY"))


def _record_row_valid(record: CanonicalRecord, findings: Sequence[Finding]) -> bool:
    for finding in findings:
        if finding.severity not in VALIDITY_SEVERITIES or finding.sheet != record.sheet: continue
        if finding.row_number == record.provenance.source_row_number: return False
        if finding.row_number is None and finding.code in SHEET_LEVEL_ROW_BLOCKING_CODES: return False
    return True


def _apply_row_validity(records: Sequence[CanonicalRecord], findings: Sequence[Finding]) -> list[CanonicalRecord]:
    return [replace(record, row_valid=_record_row_valid(record, findings), eligible_for_analysis=False, active=False) for record in records]


def _group_key(record: CanonicalRecord) -> tuple[Any, ...]:
    values = record.canonical_values
    return values.get("SOURCING_EVENT_ID"), values.get("RFQ_NUMBER"), values.get("RFQ_ITEM"), values.get("SUPPLIER_ID")


def _activate_latest_valid_versions(records: Sequence[CanonicalRecord], findings: list[Finding]) -> list[CanonicalRecord]:
    groups: dict[tuple[Any, ...], list[CanonicalRecord]] = {}
    for record in records: groups.setdefault(_group_key(record), []).append(record)
    result: list[CanonicalRecord] = []
    for key, group in groups.items():
        valid = [item for item in group if item.row_valid and isinstance(item.canonical_values.get("QUOTATION_VERSION"), int)]
        if not valid: result.extend(replace(item, active=False, eligible_for_analysis=False) for item in group); continue
        latest = max(item.canonical_values["QUOTATION_VERSION"] for item in valid); winners = [item for item in valid if item.canonical_values.get("QUOTATION_VERSION") == latest]
        if len(winners) != 1:
            findings.append(Finding("Blocking", "QUOTATION_VERSION_CONFLICT", f"Highest valid quotation version is ambiguous for group {key}.", "RFQ_QUOTES")); result.extend(replace(item, active=False, eligible_for_analysis=False) for item in group); continue
        winner = winners[0]; result.extend(replace(item, active=item is winner, eligible_for_analysis=item is winner and item.row_valid) for item in group)
    return result


def _apply_event_eligibility(records: Sequence[CanonicalRecord], selected: str | None, required: bool) -> list[CanonicalRecord]:
    if required: return [replace(item, eligible_for_analysis=False) for item in records]
    if selected is None: return list(records)
    return [replace(item, eligible_for_analysis=item.eligible_for_analysis and str(item.canonical_values.get("SOURCING_EVENT_ID")) == selected) for item in records]


def _validate_no_valid_rfq(records: Sequence[CanonicalRecord], findings: list[Finding], *, selected_event: str | None) -> None:
    if not records: findings.append(Finding("Fatal", "RFQ_QUOTES_EMPTY", "RFQ_QUOTES must contain at least one non-empty quotation row.", "RFQ_QUOTES")); return
    if not any(item.active and item.eligible_for_analysis for item in records):
        code = "NO_VALID_SELECTED_EVENT_QUOTATIONS" if selected_event is not None else "NO_VALID_QUOTATION_RECORDS"
        findings.append(Finding("Fatal", code, "No valid active quotation records remain after adapter validation.", "RFQ_QUOTES"))


def _validate_supplier_counts(records: Sequence[CanonicalRecord], findings: list[Finding]) -> None:
    suppliers: dict[tuple[Any, Any], set[Any]] = {}
    for item in records:
        if item.active and item.eligible_for_analysis:
            values = item.canonical_values; suppliers.setdefault((values.get("RFQ_NUMBER"), values.get("RFQ_ITEM")), set()).add(values.get("SUPPLIER_ID"))
    for key, values in suppliers.items():
        if len({value for value in values if value is not None}) < 2: findings.append(Finding("Blocking", "MINIMUM_SUPPLIER_COUNT_NOT_MET", f"RFQ item {key} has fewer than two valid suppliers.", "RFQ_QUOTES"))


def _json_value(value: Any) -> Any:
    if isinstance(value, Decimal): return int(value) if value == value.to_integral_value() else float(value)
    if isinstance(value, datetime): return value.isoformat()
    if isinstance(value, date): return value.isoformat()
    if isinstance(value, Mapping): return {str(key): _json_value(child) for key, child in value.items() if child is not None}
    if isinstance(value, (list, tuple)): return [_json_value(child) for child in value]
    return value


def _runtime_schema_validation(bundle: Any, rfq: Sequence[CanonicalRecord], po: Sequence[CanonicalRecord], metadata: Mapping[str, Any] | None, ranking: Sequence[CanonicalRankingRecord], findings: list[Finding]) -> set[int]:
    if bundle.validator is None: return set()
    instance = {
        "RFQ_QUOTES": [_json_value(item.canonical_values) for item in rfq],
        "PO_HISTORY": [_json_value(item.canonical_values) for item in po],
        "UPLOAD_METADATA": [] if metadata is None else [_json_value(metadata)],
        "SUPPLIER_RANKING_INPUTS": [_json_value(item.canonical_values) for item in ranking],
    }
    invalid_rows: set[int] = set()
    for error in sorted(bundle.validator.iter_errors(instance), key=lambda item: list(item.absolute_path)):
        path = list(error.absolute_path); sheet = str(path[0]) if path else None; row_number = None; field = None
        if sheet == RANKING_SHEET and len(path) > 1 and isinstance(path[1], int):
            row_number = int(path[1]) + 2; invalid_rows.add(row_number)
        if path and isinstance(path[-1], str): field = str(path[-1])
        severity = "Blocking" if sheet == RANKING_SHEET else "Fatal"
        findings.append(Finding(severity, "RANKING_ROW_SCHEMA_INVALID" if sheet == RANKING_SHEET else "WORKBOOK_SCHEMA_INVALID", error.message, sheet, row_number, field))
    return invalid_rows


def adapt_v13_workbook(source: str | Path | bytes | bytearray | BinaryIO, *, filename: str | None = None, selected_sourcing_event_id: str | None = None, confirmed_mappings: Iterable[tuple[str, str, str]] = (), ranking_confirmations: Iterable[RankingMappingConfirmation] = (), evaluation_date: date | None = None, max_file_size_bytes: int = DEFAULT_MAX_FILE_SIZE_BYTES, schema_path: Path = SCHEMA_PATH, alias_path: Path = ALIAS_PATH) -> AdapterResult:
    """Return typed adapter and ranking evidence without invoking downstream engines."""
    resolved_filename = _filename(source, filename)
    try: load_erp_workbook(source, filename=resolved_filename, max_file_size_bytes=max_file_size_bytes)
    except WorkbookLoadError as exc: raise WorkbookAdapterError(str(exc)) from exc
    payload = _read_bytes(source); upload_hash = sha256(payload).hexdigest(); findings: list[Finding] = []; confirmed = set(confirmed_mappings)
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=False, keep_links=False)
    ranking_records: list[CanonicalRankingRecord] = []; ranking_reviews: list[MappingReview] = []; confirmed_origins: set[tuple[str, str]] = set(); pending_suppliers: set[str] = set(); bundle = None
    try:
        version = _detect_schema_version(workbook)
        if schema_path != SCHEMA_PATH or alias_path != ALIAS_PATH:
            schema = _load_contract(schema_path); registry = _load_contract(alias_path); approved_sheets = APPROVED_SHEETS; core_schema = schema
        else:
            try: bundle = load_contract_bundle(version)
            except RankingContractError as exc:
                if version not in {"1.3.0", "1.3.1"}:
                    findings.append(Finding("Fatal", "SCHEMA_VERSION_UNSUPPORTED", str(exc), "UPLOAD_METADATA", 2, "SCHEMA_VERSION")); bundle = load_contract_bundle("1.3.0")
                else: raise WorkbookAdapterError(str(exc)) from exc
            schema = bundle.schema; registry = bundle.alias_registry; approved_sheets = bundle.approved_sheets; core_schema = bundle.core_schema
        for unknown in sorted(set(workbook.sheetnames) - set(approved_sheets)): findings.append(Finding("Information", "UNKNOWN_SHEET_IGNORED", f"Unknown sheet '{unknown}' is not interpreted.", unknown))
        if "RFQ_QUOTES" not in workbook.sheetnames: findings.append(Finding("Fatal", "RFQ_QUOTES_MISSING", "Mandatory sheet 'RFQ_QUOTES' is missing."))
        metadata_records, metadata_reviews = _parse_sheet(workbook, "UPLOAD_METADATA", resolved_filename, upload_hash, None, core_schema, registry, confirmed, findings)
        metadata = _validate_metadata(metadata_records, core_schema, findings)
        if version == "1.3.1" and metadata is not None:
            metadata = dict(metadata); metadata["SCHEMA_VERSION"] = "1.3.1"; findings[:] = [item for item in findings if not (item.code == "SCHEMA_VERSION_UNSUPPORTED" and item.field_name == "SCHEMA_VERSION")]
        source_hash = None if metadata is None else metadata.get("SOURCE_FILE_HASH_SHA256")
        inferred_mode = "FULL_SOURCING_REVIEW" if "PO_HISTORY" in workbook.sheetnames else "QUICK_RFQ"; mode = str((metadata or {}).get("UPLOAD_MODE") or inferred_mode)
        rfq_records, rfq_reviews = _parse_sheet(workbook, "RFQ_QUOTES", resolved_filename, upload_hash, source_hash, core_schema, registry, confirmed, findings)
        po_records, po_reviews = _parse_sheet(workbook, "PO_HISTORY", resolved_filename, upload_hash, source_hash, core_schema, registry, confirmed, findings)
        if version == "1.3.1": ranking_records, ranking_reviews, confirmed_origins, pending_suppliers = _parse_ranking_sheet(workbook, resolved_filename, upload_hash, source_hash, schema, registry, tuple(ranking_confirmations), findings)
    finally: workbook.close()
    if mode == "FULL_SOURCING_REVIEW" and not po_records: findings.append(Finding("Warning", "PO_HISTORY_UNAVAILABLE", "Full review has no valid PO history rows.", "PO_HISTORY"))
    _validate_values([*rfq_records, *po_records], findings); _validate_row_currencies([*rfq_records, *po_records], core_schema, findings); _validate_keys([*rfq_records, *po_records], findings)
    rfq_records = _apply_row_validity(rfq_records, findings); po_records = _apply_row_validity(po_records, findings); rfq_records = _activate_latest_valid_versions(rfq_records, findings)
    event_ids = tuple(sorted({str(item.canonical_values["SOURCING_EVENT_ID"]) for item in rfq_records if item.canonical_values.get("SOURCING_EVENT_ID") is not None})); selection_required = len(event_ids) > 1 and selected_sourcing_event_id is None
    if selection_required: findings.append(Finding("Blocking", "SOURCING_EVENT_SELECTION_REQUIRED", "Multiple sourcing events require explicit selection.", "RFQ_QUOTES"))
    if selected_sourcing_event_id is not None and selected_sourcing_event_id not in event_ids: findings.append(Finding("Fatal", "SOURCING_EVENT_SELECTION_INVALID", f"Selected sourcing event '{selected_sourcing_event_id}' is not present.", "RFQ_QUOTES"))
    if selected_sourcing_event_id is not None: rfq_records = [item for item in rfq_records if str(item.canonical_values.get("SOURCING_EVENT_ID")) == selected_sourcing_event_id]
    rfq_records = _apply_event_eligibility(rfq_records, selected_sourcing_event_id, selection_required); _validate_no_valid_rfq(rfq_records, findings, selected_event=selected_sourcing_event_id); _validate_supplier_counts(rfq_records, findings)
    ranking_evidence: tuple[CanonicalFieldEvidenceResult, ...] = (); ranking_matches: tuple[RankingScopeMatch, ...] = (); ranking_eligibility: tuple[RankingModeEligibility, ...] = ()
    if version == "1.3.1":
        schema_invalid_rows = _runtime_schema_validation(bundle, rfq_records, po_records, metadata, ranking_records, findings) if bundle is not None else set()
        if schema_invalid_rows: ranking_records = [replace(record, row_valid=False) if record.provenance.source_row_number in schema_invalid_rows else record for record in ranking_records]
        findings.extend(cross_row_findings(ranking_records, Finding))
        ranking_evidence = generate_evidence_results(ranking_records, evaluation_date or date.today(), Finding, confirmed_origins=confirmed_origins, schema_invalid_rows=schema_invalid_rows)
        for item in ranking_evidence: findings.extend(item.validation_findings)
        ranking_matches = match_ranking_records(rfq_records, ranking_records, ranking_evidence, Finding)
        for match in ranking_matches:
            findings.extend(match.blocking_findings)
        ranking_eligibility = calculate_mode_eligibility(mode, ranking_matches, ranking_evidence, Finding, pending_suppliers)
        for eligibility in ranking_eligibility: findings.extend(eligibility.blocking_findings)
        if ranking_eligibility and all(item.status == "RANKING_REVIEW_COMPLETE" for item in ranking_eligibility): findings.append(Finding("Information", "CANONICAL_RANKING_INPUTS_AVAILABLE_FOR_REVIEW", "Canonical ranking inputs are available for governed review only.", RANKING_SHEET))
    return AdapterResult(resolved_filename, mode, version, str(registry.get("registry_version", "unknown")), upload_hash, source_hash, selected_sourcing_event_id, event_ids, tuple(rfq_records), tuple(po_records), metadata, tuple([*metadata_reviews, *rfq_reviews, *po_reviews, *ranking_reviews]), tuple(findings), tuple(ranking_records), ranking_evidence, ranking_matches, ranking_eligibility)
