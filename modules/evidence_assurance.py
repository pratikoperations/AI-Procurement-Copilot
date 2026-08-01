"""Runtime assurance for registered Excel sheets and JSON evidence paths."""
from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
import json
from typing import Any, Mapping

from openpyxl import load_workbook

from modules.export_evidence_registry import EXPORT_EVIDENCE, STEEL_EXCEL_LOCATIONS


@dataclass(frozen=True)
class EvidenceCheckResult:
    evidence_id: str
    export_type: str
    expected_locations: tuple[str, ...]
    present_locations: tuple[str, ...]
    missing_locations: tuple[str, ...]
    audience: str
    schema_change: bool
    classification: str
    blocking_status: str
    human_review_status: str = "required"


def _locations(location: str) -> tuple[str, ...]:
    return tuple(part.strip() for part in location.split(";") if part.strip())


def _json_path_exists(payload: Any, path: str) -> bool:
    current = payload
    for segment in path.split("."):
        if isinstance(current, Mapping) and segment in current:
            current = current[segment]
        else:
            return False
    return True


def assure_excel_evidence(workbook_bytes: bytes, evidence_id: str) -> EvidenceCheckResult:
    evidence = next(item for item in EXPORT_EVIDENCE if item.evidence_id == evidence_id)
    if evidence.export_type != "excel":
        raise ValueError(f"{evidence_id} is not Excel evidence")
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    present_sheets = tuple(workbook.sheetnames)
    expected = _locations(evidence.location)
    present = tuple(location for location in expected if location in present_sheets)
    missing = tuple(location for location in expected if location not in present_sheets)
    return EvidenceCheckResult(
        evidence_id=evidence.evidence_id,
        export_type=evidence.export_type,
        expected_locations=expected,
        present_locations=present,
        missing_locations=missing,
        audience=evidence.audience,
        schema_change=evidence.schema_change,
        classification="exact_match" if not missing else "export_path_inconsistency",
        blocking_status="clear" if not missing else "blocked",
    )


def assure_json_evidence(payload: str | bytes | Mapping[str, Any], evidence_id: str) -> EvidenceCheckResult:
    evidence = next(item for item in EXPORT_EVIDENCE if item.evidence_id == evidence_id)
    if evidence.export_type != "json":
        raise ValueError(f"{evidence_id} is not JSON evidence")
    if isinstance(payload, bytes):
        payload = payload.decode("utf-8")
    if isinstance(payload, str):
        payload = json.loads(payload)
    expected = _locations(evidence.location)
    present = tuple(path for path in expected if _json_path_exists(payload, path))
    missing = tuple(path for path in expected if path not in present)
    return EvidenceCheckResult(
        evidence_id=evidence.evidence_id,
        export_type=evidence.export_type,
        expected_locations=expected,
        present_locations=present,
        missing_locations=missing,
        audience=evidence.audience,
        schema_change=evidence.schema_change,
        classification="exact_match" if not missing else "export_path_inconsistency",
        blocking_status="clear" if not missing else "blocked",
    )


def steel_sheet_contract() -> tuple[str, ...]:
    """Expose the exact registered Steel workbook contract without changing it."""
    return tuple(STEEL_EXCEL_LOCATIONS)
