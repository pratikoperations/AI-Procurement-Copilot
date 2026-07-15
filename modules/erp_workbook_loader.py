"""Safe structural loader for Version 1.1 ERP `.xlsx` workbooks.

The loader performs file-level and package-safety checks and extracts sheet
metadata only. It does not normalize, score, persist, or execute workbook
content.
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import BinaryIO, Mapping
from zipfile import BadZipFile, ZipFile, is_zipfile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


DEFAULT_MAX_FILE_SIZE_BYTES = 25 * 1024 * 1024
_MACRO_SUFFIX = "vbaproject.bin"
_EXTERNAL_LINK_PREFIX = "xl/externallinks/"
_QUERY_TABLE_PREFIX = "xl/querytables/"
_CONNECTION_PARTS = {"xl/connections.xml"}
_EXTERNAL_DATA_PREFIXES = (
    "xl/externaldata/",
    "xl/model/",
)


class WorkbookLoadError(ValueError):
    """Raised when an uploaded workbook cannot be accepted safely."""


@dataclass(frozen=True)
class SheetSummary:
    """Read-only structural metadata for one workbook sheet."""

    name: str
    headers: tuple[object, ...]
    row_count: int
    column_count: int
    is_empty: bool


@dataclass(frozen=True)
class WorkbookSummary:
    """Read-only structural summary returned by the loader."""

    filename: str
    file_size_bytes: int
    sheet_names: tuple[str, ...]
    sheets: Mapping[str, SheetSummary]


def _resolve_filename(source: object, filename: str | None) -> str:
    if filename:
        return Path(filename).name
    if isinstance(source, (str, Path)):
        return Path(source).name
    source_name = getattr(source, "name", None)
    if isinstance(source_name, str) and source_name:
        return Path(source_name).name
    raise WorkbookLoadError("A filename is required to validate the workbook type.")


def _read_source_bytes(source: str | Path | bytes | bytearray | BinaryIO) -> bytes:
    if isinstance(source, (str, Path)):
        path = Path(source)
        if not path.is_file():
            raise WorkbookLoadError(f"Workbook file not found: {path}")
        try:
            return path.read_bytes()
        except OSError as exc:
            raise WorkbookLoadError(f"Workbook could not be read: {exc}") from exc

    if isinstance(source, (bytes, bytearray)):
        return bytes(source)

    read = getattr(source, "read", None)
    if not callable(read):
        raise WorkbookLoadError("Workbook source must be a path, bytes, or binary file object.")

    original_position = None
    tell = getattr(source, "tell", None)
    seek = getattr(source, "seek", None)
    try:
        if callable(tell):
            original_position = tell()
        if callable(seek):
            seek(0)
        payload = read()
    except (OSError, ValueError) as exc:
        raise WorkbookLoadError(f"Workbook could not be read: {exc}") from exc
    finally:
        if original_position is not None and callable(seek):
            try:
                seek(original_position)
            except (OSError, ValueError):
                pass

    if not isinstance(payload, (bytes, bytearray)):
        raise WorkbookLoadError("Workbook file object must return binary bytes.")
    return bytes(payload)


def _validate_package_members(names: set[str], filename: str) -> None:
    lowered = {name.lower() for name in names}

    if any(name.endswith(_MACRO_SUFFIX) for name in lowered):
        raise WorkbookLoadError("Macro-enabled workbooks are not permitted.")
    if any(name.startswith(_EXTERNAL_LINK_PREFIX) for name in lowered):
        raise WorkbookLoadError(
            f"Workbook '{filename}' contains external workbook links, which are not permitted."
        )
    if lowered.intersection(_CONNECTION_PARTS):
        raise WorkbookLoadError(
            f"Workbook '{filename}' contains connection definitions, which are not permitted."
        )
    if any(name.startswith(_QUERY_TABLE_PREFIX) for name in lowered):
        raise WorkbookLoadError(
            f"Workbook '{filename}' contains query-table definitions, which are not permitted."
        )
    if any(
        name.startswith(prefix)
        for name in lowered
        for prefix in _EXTERNAL_DATA_PREFIXES
    ):
        raise WorkbookLoadError(
            f"Workbook '{filename}' contains external-data definitions, which are not permitted."
        )


def _validate_xlsx_container(payload: bytes, filename: str) -> None:
    if not is_zipfile(BytesIO(payload)):
        raise WorkbookLoadError(
            f"'{filename}' is not a valid XLSX package; renamed or corrupt files are rejected."
        )

    try:
        with ZipFile(BytesIO(payload)) as archive:
            names = set(archive.namelist())
            required_parts = {"[Content_Types].xml", "xl/workbook.xml"}
            if not required_parts.issubset(names):
                raise WorkbookLoadError(
                    f"'{filename}' is missing required XLSX package components."
                )
            _validate_package_members(names, filename)
    except BadZipFile as exc:
        raise WorkbookLoadError(f"'{filename}' is a corrupt XLSX package.") from exc


def _extract_sheet_summary(worksheet: object) -> SheetSummary:
    rows = worksheet.iter_rows(values_only=True)
    first_row = next(rows, None)
    headers = tuple(first_row or ())

    non_empty_rows = 0
    for row in rows:
        if any(value is not None and str(value).strip() != "" for value in row):
            non_empty_rows += 1

    return SheetSummary(
        name=worksheet.title,
        headers=headers,
        row_count=non_empty_rows,
        column_count=len(headers),
        is_empty=non_empty_rows == 0,
    )


def load_erp_workbook(
    source: str | Path | bytes | bytearray | BinaryIO,
    *,
    filename: str | None = None,
    max_file_size_bytes: int = DEFAULT_MAX_FILE_SIZE_BYTES,
) -> WorkbookSummary:
    """Validate and structurally inspect one `.xlsx` workbook.

    Formulas are read as stored expressions and are never executed. The workbook
    is opened read-only, no extracted data is persisted, and callers receive only
    immutable structural metadata.
    """

    resolved_filename = _resolve_filename(source, filename)
    if Path(resolved_filename).suffix.lower() != ".xlsx":
        raise WorkbookLoadError("Only non-macro `.xlsx` workbooks are accepted.")
    if max_file_size_bytes <= 0:
        raise WorkbookLoadError("Maximum file size must be greater than zero.")

    payload = _read_source_bytes(source)
    if not payload:
        raise WorkbookLoadError("Workbook is empty.")
    if len(payload) > max_file_size_bytes:
        raise WorkbookLoadError(
            f"Workbook exceeds the {max_file_size_bytes}-byte upload limit."
        )

    _validate_xlsx_container(payload, resolved_filename)

    try:
        workbook = load_workbook(
            BytesIO(payload),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as exc:
        raise WorkbookLoadError(
            f"Workbook '{resolved_filename}' could not be opened safely: {exc}"
        ) from exc

    try:
        summaries = {
            worksheet.title: _extract_sheet_summary(worksheet)
            for worksheet in workbook.worksheets
        }
    finally:
        workbook.close()

    if not summaries:
        raise WorkbookLoadError("Workbook must contain at least one worksheet.")

    return WorkbookSummary(
        filename=resolved_filename,
        file_size_bytes=len(payload),
        sheet_names=tuple(summaries),
        sheets=summaries,
    )
